import os

from services.onedrive_service import OneDriveService
from services.user_master_reader import UserMasterReader


def main():
    service = OneDriveService()

    config = service.auth  # 認証初期化確認用

    user_master_url = (
        "https://otameshisharepoint-my.sharepoint.com/"
        ":x:/g/personal/"
        "teru_toshimori_otameshisharepoint_onmicrosoft_com/"
        "IQAxlcofs-vSQa1DDIJGCJaVAScZ7ABTk1wM6al5WjR6E7w"
        "?e=0cG5xj"
    )

    temp_path = None

    try:
        # URLからファイル情報を取得
        item = service.resolve_shared_item(user_master_url)

        print("解決したファイル名:", item["file_name"])
        print("Drive ID:", item["drive_id"])
        print("Item ID:", item["item_id"])

        # 一時ダウンロード
        temp_path, _ = service.download_to_temp(user_master_url)

        print("一時ファイル:", temp_path)
        print("サイズ:", os.path.getsize(temp_path), "bytes")

        # Graphのログインユーザー情報
        user = service.get_current_user()

        print("Graph mail:", repr(user.get("mail")))
        print(
            "Graph userPrincipalName:",
            repr(user.get("userPrincipalName")),
        )

        email_candidates = [
            user.get("mail") or "",
            user.get("userPrincipalName") or "",
        ]

        reader = UserMasterReader()

        from openpyxl import load_workbook

        wb = load_workbook(temp_path)

        for sheet in wb.worksheets:

            print("======", sheet.title, "======")

            for row in sheet.iter_rows():

                print(
                    row[0].value,
                    row[1].value,
                    row[2].value,
                )

        wb.close()

    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


if __name__ == "__main__":
    main()