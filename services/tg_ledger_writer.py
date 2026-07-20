import os
import re
import logging
from copy import copy
from datetime import datetime

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class TgLedgerWriter:
    """
    TG台帳記入クラス

    処理の流れ

        台帳Excelを開く
            ↓
        PDF解析結果を1件ずつ処理
            ↓
        台帳最終行を取得
            ↓
        新しい行へ書式コピー
            ↓
        OCR結果を書き込む
            ↓
        採番した見積番号を保持
            ↓
        台帳保存
    """

    def write(
        self,
        ledger_path: str,
        pdf_results: list
    ):

        logger.info("========== 台帳記入開始 ==========")

        # ---------------------------------
        # 台帳Excelを開く
        # ---------------------------------
        wb = load_workbook(ledger_path)

        # 豊田合成シート取得
        ws = wb["豊田合成"]

        # ---------------------------------
        # PDFごとに処理
        # ---------------------------------
        for item in pdf_results:

            data = item["data"]

            # ---------------------------------
            # PDFファイル名取得
            # ---------------------------------
            pdf_name = os.path.basename(item["pdf_path"])

            # 「仕様書_」と「.pdf」を除いた文字列を取得
            # 例)
            # 仕様書_QY6247.pdf
            #      ↓
            # QY6247
            match = re.search(
                r"仕様書_(.+?)\.pdf",
                pdf_name,
                re.IGNORECASE
            )

            if match:
                pdf_name = match.group(1)

            # ---------------------------------
            # E列の最終入力行を取得
            # ---------------------------------
            last_row = 1

            # 下から上へ検索し、
            # 最後に値が入っているE列を探す
            for r in range(ws.max_row, 1, -1):

                if ws[f"E{r}"].value:

                    last_row = r
                    break

            # 新しく追加する行
            new_row = last_row + 1

            # ---------------------------------
            # A～K列の書式コピー
            # ---------------------------------
            # 上の行と同じ書式になるようコピー
            for col in range(1, 12):

                src = ws.cell(
                    row=last_row,
                    column=col
                )

                dst = ws.cell(
                    row=new_row,
                    column=col
                )

                if src.has_style:
                    dst._style = copy(src._style)

                if src.number_format:
                    dst.number_format = copy(src.number_format)

                if src.font:
                    dst.font = copy(src.font)

                if src.fill:
                    dst.fill = copy(src.fill)

                if src.border:
                    dst.border = copy(src.border)

                if src.alignment:
                    dst.alignment = copy(src.alignment)

                if src.protection:
                    dst.protection = copy(src.protection)

            # ---------------------------------
            # B列
            # No.
            # ---------------------------------
            last_b = ws[f"B{last_row}"].value or 0

            ws[f"B{new_row}"] = int(last_b) + 1

            # ---------------------------------
            # D列
            # 見積番号
            # ---------------------------------
            last_d = ws[f"D{last_row}"].value or 0

            ws[f"D{new_row}"] = int(last_d) + 1

            # ---------------------------------
            # E列
            # PDF名
            # ---------------------------------
            ws[f"E{new_row}"] = pdf_name

            # ---------------------------------
            # G列
            # OCRで取得した品名
            # 「委託」は削除する
            # ---------------------------------
            subject = item["data"].get("品名", "")

            subject = re.sub(
                r"委託$",
                "",
                subject
            ).strip()

            ws[f"G{new_row}"] = subject

            # ---------------------------------
            # H列
            # OCRで取得した金額
            # ---------------------------------
            ws[f"H{new_row}"] = item["data"].get(
                "金額",
                ""
            )

            # ---------------------------------
            # I列
            # 翌月を「○月」で入力
            # 例)
            # 2026/07/17 → 8月
            # ---------------------------------
            today = datetime.today()

            next_month = today + relativedelta(
                months=1
            )

            ws[f"I{new_row}"] = f"{next_month.month}月"

            # ---------------------------------
            # J列
            # 今日の日付
            # ---------------------------------
            ws[f"J{new_row}"] = today.strftime(
                "%Y/%m/%d"
            )

            # ---------------------------------
            # 採番した見積番号を保持
            # 後続のExcel出力(G1)で使用する
            # ---------------------------------
            estimate_no = ws[f"D{new_row}"].value

            item["estimate_no"] = estimate_no

            logger.info(
                "保持した見積番号 : %s",
                estimate_no
            )

        # ---------------------------------
        # 台帳保存
        # ---------------------------------
        wb.save(ledger_path)

        logger.info("========== 台帳記入終了 ==========")