from copy import copy
from datetime import datetime
import getpass
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class LedgerWriter:
    """
    管理台帳へ新しい行を追加する。
    """

    def write(
        self,
        excel_path: str,
        data,
        issuer_name: str = "",
    ) -> dict:
        """
        管理台帳へ1行追加する。

        Args:
            excel_path:
                一時ダウンロードした管理台帳のパス

            data:
                EstimateData

            issuer_name:
                Microsoft Graphから取得した
                メールアドレスまたはアカウント名
        """

        workbook = load_workbook(
            excel_path,
            keep_links=False,
        )

        sheet = None

        try:
            sheet = self.find_sheet_by_department(
                workbook,
                data.department,
            )

            if sheet is None:
                raise ValueError(
                    "依頼部署に該当するシートが"
                    "見つかりません。\n"
                    f"依頼部署：{data.department}"
                )

            last_row = self.find_last_row(sheet)
            new_row = last_row + 1

            previous_no = sheet[
                f"B{last_row}"
            ].value

            previous_estimate_no = sheet[
                f"D{last_row}"
            ].value

            new_no = (
                self.to_int(previous_no) + 1
            )

            new_estimate_no = (
                self.to_int(
                    previous_estimate_no
                ) + 1
            )

            today = datetime.today()

            estimate_month = (
                self.get_next_month_text(today)
            )

            # Microsoftアカウント情報が空なら、
            # Windowsログインユーザー名を使用
            user_name = (
                issuer_name.strip()
                if issuer_name
                and issuer_name.strip()
                else getpass.getuser()
            )

            # ITK14642 → 14642
            application_number = self.to_int(
                data.application_no
            )

            # 前行の書式を新しい行へコピー
            self.copy_row_style(
                sheet=sheet,
                source_row=last_row,
                target_row=new_row,
                start_column=2,
                end_column=12,
            )

            # =============================
            # 値の書き込み
            # =============================
            sheet[f"B{new_row}"] = new_no

            sheet[f"C{new_row}"] = (
                data.department
            )

            sheet[f"D{new_row}"] = (
                new_estimate_no
            )

            sheet[f"F{new_row}"] = (
                application_number
            )

            sheet[f"G{new_row}"] = (
                data.model_code
            )

            sheet[f"I{new_row}"] = (
                estimate_month
            )

            # 日付のみを書き込む
            sheet[f"J{new_row}"] = (
                today.date()
            )

            # Microsoftアカウントのメール等
            sheet[f"K{new_row}"] = user_name

            # =============================
            # フォント
            # B列～L列
            # =============================
            for column in range(2, 13):

                cell = sheet.cell(
                    row=new_row,
                    column=column,
                )

                current_font = cell.font

                cell.font = Font(
                    name="Meiryo UI",
                    size=10,
                    bold=current_font.bold,
                    italic=current_font.italic,
                    vertAlign=(
                        current_font.vertAlign
                    ),
                    underline=(
                        current_font.underline
                    ),
                    strike=current_font.strike,
                    color=copy(
                        current_font.color
                    ),
                )

            # =============================
            # 配置
            # =============================

            # C列・G列は左右中央揃えを解除し、
            # 前行の左右配置を保持する
            for column_letter in ("C", "G"):

                cell = sheet[
                    f"{column_letter}{new_row}"
                ]

                current_alignment = (
                    cell.alignment
                )

                cell.alignment = Alignment(
                    horizontal=(
                        current_alignment.horizontal
                    ),
                    vertical="center",
                    wrap_text=(
                        current_alignment.wrap_text
                    ),
                    shrink_to_fit=(
                        current_alignment.shrink_to_fit
                    ),
                    text_rotation=(
                        current_alignment.text_rotation
                    ),
                )

            # その他の入力セルは中央揃え
            for column_letter in (
                "B",
                "D",
                "F",
                "I",
                "J",
                "K",
            ):

                cell = sheet[
                    f"{column_letter}{new_row}"
                ]

                current_alignment = (
                    cell.alignment
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=(
                        current_alignment.wrap_text
                    ),
                    shrink_to_fit=(
                        current_alignment.shrink_to_fit
                    ),
                    text_rotation=(
                        current_alignment.text_rotation
                    ),
                )

            # =============================
            # 表示形式
            # =============================

            # F列：数値のまま保持し、
            # Excel表示時にITKを付ける
            sheet[
                f"F{new_row}"
            ].number_format = '"ITK"0'

            # J列：日付のみ表示
            sheet[
                f"J{new_row}"
            ].number_format = "yyyy/mm/dd"

            # =============================
            # B列幅
            # =============================
            self.adjust_column_width(
                sheet=sheet,
                column_letter="B",
                minimum_width=4,
                maximum_width=12,
            )

            workbook.save(excel_path)

            return {
                "sheet_name": sheet.title,
                "row": new_row,
                "estimate_no": str(
                    new_estimate_no
                ),
                "issue_date": today.strftime(
                    "%Y/%m/%d"
                ),
                "user_name": user_name,
            }

        finally:
            workbook.close()

    # =====================================
    # 全シートのC列から依頼部署を検索
    # =====================================
    def find_sheet_by_department(
        self,
        workbook,
        department: str,
    ):

        normalized_department = (
            self.normalize_text(department)
        )

        if not normalized_department:
            return None

        for sheet in workbook.worksheets:

            for cell in sheet["C"]:

                if cell.value is None:
                    continue

                normalized_value = (
                    self.normalize_text(
                        cell.value
                    )
                )

                if (
                    normalized_value
                    == normalized_department
                    or normalized_department
                    in normalized_value
                    or normalized_value
                    in normalized_department
                ):
                    return sheet

        return None

    # =====================================
    # 文字列正規化
    # =====================================
    def normalize_text(
        self,
        text,
    ) -> str:

        if text is None:
            return ""

        normalized = unicodedata.normalize(
            "NFKC",
            str(text).strip(),
        )

        normalized = normalized.replace(
            "　",
            "",
        )

        normalized = normalized.replace(
            " ",
            "",
        )

        return normalized

    # =====================================
    # B列の最終データ行を取得
    # =====================================
    def find_last_row(
        self,
        sheet,
    ) -> int:

        for row_number in range(
            sheet.max_row,
            1,
            -1,
        ):

            if (
                sheet[
                    f"B{row_number}"
                ].value
                is not None
            ):
                return row_number

        return 1

    # =====================================
    # 数値変換
    # =====================================
    def to_int(
        self,
        value,
    ) -> int:

        if value is None:
            return 0

        text = str(value).strip()

        try:
            return int(text)

        except (TypeError, ValueError):

            digits = "".join(
                character
                for character in text
                if character.isdigit()
            )

            return (
                int(digits)
                if digits
                else 0
            )

    # =====================================
    # 翌月を「8月」形式で返す
    # =====================================
    def get_next_month_text(
        self,
        date_value: datetime,
    ) -> str:

        next_month = (
            date_value.month + 1
        )

        if next_month == 13:
            next_month = 1

        return f"{next_month}月"

    # =====================================
    # 前行の書式を新しい行へコピー
    # =====================================
    def copy_row_style(
        self,
        sheet,
        source_row: int,
        target_row: int,
        start_column: int,
        end_column: int,
    ) -> None:

        source_height = (
            sheet.row_dimensions[
                source_row
            ].height
        )

        if source_height is not None:
            sheet.row_dimensions[
                target_row
            ].height = source_height

        for column in range(
            start_column,
            end_column + 1,
        ):

            source_cell = sheet.cell(
                row=source_row,
                column=column,
            )

            target_cell = sheet.cell(
                row=target_row,
                column=column,
            )

            if source_cell.has_style:

                target_cell.font = copy(
                    source_cell.font
                )

                target_cell.fill = copy(
                    source_cell.fill
                )

                target_cell.border = copy(
                    source_cell.border
                )

                target_cell.alignment = copy(
                    source_cell.alignment
                )

                target_cell.number_format = (
                    source_cell.number_format
                )

                target_cell.protection = copy(
                    source_cell.protection
                )

    # =====================================
    # 列幅調整
    # =====================================
    def adjust_column_width(
        self,
        sheet,
        column_letter: str,
        minimum_width: float = 4,
        maximum_width: float = 50,
    ) -> None:

        maximum_length = 0

        for cell in sheet[column_letter]:

            if cell.value is None:
                continue

            text_length = len(
                str(cell.value)
            )

            if text_length > maximum_length:
                maximum_length = text_length

        calculated_width = (
            maximum_length + 2
        )

        calculated_width = max(
            minimum_width,
            calculated_width,
        )

        calculated_width = min(
            maximum_width,
            calculated_width,
        )

        sheet.column_dimensions[
            column_letter
        ].width = calculated_width