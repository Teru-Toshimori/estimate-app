import os
import shutil
import logging
from datetime import datetime

import openpyxl
import win32com.client

logger = logging.getLogger(__name__)


class TgExcelWriter:
    """
    TG見積書 Excel出力クラス

    処理の流れ

        元Excelをコピー
            ↓
        VBAを保持したまま開く
            ↓
        OCR結果・固定値を書き込む
            ↓
        保存
            ↓
        PDFへ変換
    """

    def write(
        self,
        input_excel_path: str,
        output_excel_path: str,
        data: dict,
        estimate_no,
    ):
        """
        Excel出力

        Parameters
        ----------
        input_excel_path
            入力フォルダにある元Excel

        output_excel_path
            出力先

        data
            OCR結果

        estimate_no
            台帳で採番した見積番号
        """

        logger.info("========== Excel出力開始 ==========")

        # -------------------------------------------------
        # 元となるExcelファイルが存在するか確認
        # -------------------------------------------------
        if not os.path.exists(input_excel_path):
            raise FileNotFoundError(input_excel_path)

        # -------------------------------------------------
        # 元Excelを出力先へコピー
        # （元ファイルは変更しない）
        # -------------------------------------------------
        shutil.copy2(
            input_excel_path,
            output_excel_path
        )

        logger.info(
            "Excelコピー : %s",
            output_excel_path
        )

        # -------------------------------------------------
        # VBA(マクロ)を保持したままExcelを開く
        # keep_vba=True が重要
        # -------------------------------------------------
        wb = openpyxl.load_workbook(
            output_excel_path,
            keep_vba=True
        )

        # 先頭シート取得
        ws = wb.active

        # ---------------------------------
        # G1
        # 台帳で採番した見積番号
        # ---------------------------------

        ws["G1"] = estimate_no

        logger.info(
            "G1(見積番号) : %s",
            estimate_no
        )

        # ---------------------------------
        # G2
        # 今日の日付
        # ---------------------------------

        ws["G2"] = datetime.today().strftime(
            "%Y年%m月%d日"
        )

        # ---------------------------------
        # B6
        # 元ExcelのF6をコピー
        # （F6は変更しない）
        # ---------------------------------

        ws["B6"] = ws["F6"].value

        # ---------------------------------
        # 固定値入力
        # ---------------------------------

        # 工数
        ws["B8"] = 60

        # 数量
        ws["F8"] = 1

        # 発注元会社名
        ws["F4"] = "エイム株式会社"

        # 担当者
        ws["F5"] = "齊藤　政輝"

        # 明細No
        ws["E17"] = 1

        # ---------------------------------
        # OCR結果取得
        # ---------------------------------

        subject = data.get(
            "品名",
            ""
        )

        amount = data.get(
            "金額",
            ""
        )

        # ---------------------------------
        # 品名入力
        # ---------------------------------

        if subject:

            # 件名
            ws["B4"] = subject

            # 明細名
            ws["B17"] = subject

            logger.info(
                "品名 : %s",
                subject
            )

        # ---------------------------------
        # 金額入力
        # ---------------------------------

        if amount:

            # カンマ除去して数値へ変換
            value = int(
                str(amount).replace(",", "")
            )

            # 見積金額
            ws["B10"] = value

            # 明細金額
            ws["F17"] = value

            logger.info(
                "金額 : %s",
                value
            )

        # -------------------------------------------------
        # Excel保存
        # -------------------------------------------------
        wb.save(output_excel_path)

        # ファイルを閉じる
        wb.close()

        logger.info(
            "Excel保存 : %s",
            output_excel_path
        )


    def export_pdf(
        self,
        excel_path: str,
        pdf_path: str
    ):
        """
        Excel→PDF変換

        Windows版Excel(COM)を利用して
        ExcelファイルをPDFとして保存する。
        """

        logger.info("PDF出力開始")

        # Excelアプリケーション起動
        excel = win32com.client.DispatchEx(
            "Excel.Application"
        )

        # Excel画面を表示しない
        excel.Visible = False

        # 保存確認ダイアログ等を表示しない
        excel.DisplayAlerts = False

        try:

            # Excelファイルを開く
            wb = excel.Workbooks.Open(
                os.path.abspath(excel_path)
            )

            # PDFとして保存
            wb.ExportAsFixedFormat(
                0,      # PDF形式
                os.path.abspath(pdf_path)
            )

            # 保存せず閉じる
            wb.Close(False)

            logger.info(
                "PDF保存 : %s",
                pdf_path
            )

        finally:

            # Excel終了
            excel.Quit()