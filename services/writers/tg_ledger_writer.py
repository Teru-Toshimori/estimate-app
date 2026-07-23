import os
import re
from copy import copy
from datetime import date

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook


class DuplicateRequestNoError(ValueError):
    """発注番号が管理台帳に登録済みの場合に発生する。"""


class TgLedgerWriter:
    SHEET_NAME = "豊田合成"

    def append(
        self,
        ledger_path: str,
        request_no: str,
        data: dict,
        issuer_name: str,
    ) -> dict:
        if not os.path.exists(ledger_path):
            raise FileNotFoundError(ledger_path)

        workbook = load_workbook(ledger_path)

        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                raise KeyError(
                    f"台帳に『{self.SHEET_NAME}』シートがありません。"
                )

            worksheet = workbook[self.SHEET_NAME]

            normalized_request_no = str(request_no or "").strip()
            if not normalized_request_no:
                raise ValueError("発注番号が空のため、台帳へ追加できません。")

            duplicate_row = self.find_duplicate_request_no(
                worksheet=worksheet,
                request_no=normalized_request_no,
            )

            if duplicate_row is not None:
                raise DuplicateRequestNoError(
                    f"発注番号『{normalized_request_no}』は、"
                    f"管理台帳の{duplicate_row}行目に既に登録されています。"
                )

            last_row = self.find_last_row(worksheet)
            new_row = last_row + 1

            self.copy_row_style(
                worksheet=worksheet,
                source_row=last_row,
                target_row=new_row,
            )

            last_no = self.to_int(
                worksheet[f"B{last_row}"].value
            )
            last_estimate_no = self.to_int(
                worksheet[f"D{last_row}"].value
            )
            estimate_no = last_estimate_no + 1

            subject = re.sub(
                r"委託$",
                "",
                str(data.get("品名", "") or "").strip(),
            ).strip()

            amount = self.to_int(data.get("金額", ""))
            today = date.today()
            next_month = today + relativedelta(months=1)

            worksheet[f"B{new_row}"] = last_no + 1
            worksheet[f"D{new_row}"] = estimate_no
            worksheet[f"E{new_row}"] = normalized_request_no
            worksheet[f"G{new_row}"] = subject
            worksheet[f"H{new_row}"] = amount
            worksheet[f"I{new_row}"] = f"{next_month.month}月"
            worksheet[f"J{new_row}"] = today
            worksheet[f"J{new_row}"].number_format = "'yy/m/d"
            worksheet[f"K{new_row}"] = str(issuer_name or "").strip()

            workbook.save(ledger_path)

            return {
                "row": new_row,
                "estimate_no": estimate_no,
                "issuer_name": str(issuer_name or "").strip(),
            }

        finally:
            workbook.close()

    @staticmethod
    def find_duplicate_request_no(
        worksheet,
        request_no: str,
    ) -> int | None:
        """E列から同一の発注番号を検索し、一致した行番号を返す。"""

        target = str(request_no or "").strip().casefold()

        if not target:
            return None

        for row in range(2, worksheet.max_row + 1):
            current = str(
                worksheet[f"E{row}"].value or ""
            ).strip().casefold()

            if current == target:
                return row

        return None

    @staticmethod
    def find_last_row(worksheet) -> int:
        for row in range(worksheet.max_row, 1, -1):
            if worksheet[f"E{row}"].value not in (None, ""):
                return row

        raise RuntimeError("TG台帳の最終行を取得できません。")

    @staticmethod
    def copy_row_style(
        worksheet,
        source_row: int,
        target_row: int,
    ) -> None:
        for column in range(1, 12):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)

            if source.has_style:
                target._style = copy(source._style)

            target.number_format = copy(source.number_format)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

        worksheet.row_dimensions[target_row].height = (
            worksheet.row_dimensions[source_row].height
        )

    @staticmethod
    def to_int(value) -> int:
        if value in (None, ""):
            return 0

        text = re.sub(r"[^\d-]", "", str(value))
        return int(text) if text else 0
