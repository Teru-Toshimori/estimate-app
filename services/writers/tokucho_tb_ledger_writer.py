from copy import copy
from datetime import datetime
import getpass
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class DuplicateApplicationNoError(Exception):
    """
    業務計画書Noが管理台帳へ
    すでに登録されている場合の専用エラー。
    """

    def __init__(
        self,
        application_no: str,
        sheet_name: str,
        row_number: int,
    ):
        self.application_no = application_no
        self.sheet_name = sheet_name
        self.row_number = row_number

        super().__init__(
            f"業務計画書No「{application_no}」は"
            "管理台帳に登録済みです。\n"
            f"シート：{sheet_name}\n"
            f"行番号：{row_number}"
        )


class LedgerWriter:
    """
    管理台帳へ新しい行を追加する。

    追加する内容:
        B列：No
        C列：部署名
        D列：見積／請求番号
        F列：業務計画書No
        G列：車種コード
        H列：委託金額
        I列：見積月
        J列：発行日
        K列：発行者

    採番前に、管理台帳の全シートを対象として
    F列に同じ業務計画書Noが存在しないか確認する。
    """

    def write(
        self,
        excel_path: str,
        data,
        issuer_name: str = "",
    ) -> dict:
        """
        管理台帳へ1行追加する。

        Args:
            excel_path:
                一時ダウンロードした管理台帳のパス

            data:
                EstimateDataオブジェクト

            issuer_name:
                Microsoft Graphおよび利用者リストから
                取得した利用者名

        Returns:
            採番結果や追加行情報を格納した辞書

        Raises:
            DuplicateApplicationNoError:
                業務計画書Noが登録済みの場合
        """

        workbook = load_workbook(
            excel_path,
            keep_links=False,
        )

        try:
            # =====================================
            # 業務委託計画書Noを数値へ変換
            # 帳票によって存在しないため任意項目とする。
            # ITK14642 → 14642
            # 空欄      → None
            # =====================================
            application_no_text = str(
                getattr(
                    data,
                    "application_no",
                    "",
                )
                or ""
            ).strip()

            application_number = (
                self.to_int(application_no_text)
                if application_no_text
                else None
            )

            if (
                application_no_text
                and (
                    application_number is None
                    or application_number <= 0
                )
            ):
                raise ValueError(
                    "業務委託計画書Noを数値として"
                    "取得できませんでした。\n\n"
                    f"取得値：{application_no_text}"
                )

            # =====================================
            # 業務委託計画書Noがある場合だけ重複確認
            # =====================================
            if application_number is not None:
                duplicate_info = (
                    self.find_duplicate_application_no(
                        workbook=workbook,
                        application_no=application_number,
                    )
                )

                if duplicate_info is not None:
                    duplicate_sheet, duplicate_row = (
                        duplicate_info
                    )

                    raise DuplicateApplicationNoError(
                        application_no=application_no_text,
                        sheet_name=duplicate_sheet.title,
                        row_number=duplicate_row,
                    )

            # =====================================
            # 依頼部署から対象シートを検索
            # =====================================
            sheet = self.find_sheet_by_department(
                workbook=workbook,
                department=data.department,
            )

            if sheet is None:
                raise ValueError(
                    "依頼部署に該当するシートが"
                    "見つかりません。\n\n"
                    f"依頼部署：{data.department}"
                )

            # =====================================
            # あらかじめ採番済みの未使用行を検索
            # =====================================
            available_row = self.find_available_row(
                sheet
            )

            if available_row is not None:
                new_row = available_row

                new_no = self.to_int(
                    sheet[
                        f"B{new_row}"
                    ].value
                )

                new_estimate_no = self.to_int(
                    sheet[
                        f"D{new_row}"
                    ].value
                )

                if new_no <= 0:
                    raise ValueError(
                        "未使用行のNoを"
                        "取得できませんでした。\n\n"
                        f"シート：{sheet.title}\n"
                        f"行番号：{new_row}"
                    )

                if new_estimate_no <= 0:
                    raise ValueError(
                        "未使用行の見積／請求番号を"
                        "取得できませんでした。\n\n"
                        f"シート：{sheet.title}\n"
                        f"行番号：{new_row}"
                    )

                used_preassigned_row = True

            else:
                # =====================================
                # 採番済みの未使用行がない場合は、
                # 従来どおり最終行の下へ追加する。
                # =====================================
                last_row = self.find_last_row(
                    sheet
                )

                if last_row <= 1:
                    raise ValueError(
                        "管理台帳の最終データ行を"
                        "取得できませんでした。\n\n"
                        f"シート：{sheet.title}"
                    )

                new_row = last_row + 1

                previous_no = sheet[
                    f"B{last_row}"
                ].value

                previous_estimate_no = sheet[
                    f"D{last_row}"
                ].value

                new_no = (
                    self.to_int(previous_no)
                    + 1
                )

                new_estimate_no = (
                    self.to_int(
                        previous_estimate_no
                    )
                    + 1
                )

                if new_no <= 1:
                    raise ValueError(
                        "管理台帳のNoを"
                        "採番できませんでした。\n\n"
                        f"前行の値：{previous_no}"
                    )

                if new_estimate_no <= 1:
                    raise ValueError(
                        "見積／請求番号を"
                        "採番できませんでした。\n\n"
                        f"前行の値：{previous_estimate_no}"
                    )

                used_preassigned_row = False

            # =====================================
            # 日付・見積月・発行者
            # =====================================
            today = datetime.today()

            estimate_month = (
                self.get_next_month_text(
                    today
                )
            )

            user_name = (
                issuer_name.strip()
                if (
                    issuer_name
                    and issuer_name.strip()
                )
                else getpass.getuser()
            )

            # =====================================
            # 委託金額を数値へ変換
            # =====================================
            amount_value = (
                self.normalize_amount(
                    data.amount
                )
            )

            # =====================================
            # 新しい行を追加した場合だけ、
            # 前行の書式をB列～L列へコピーする。
            # =====================================
            if not used_preassigned_row:
                self.copy_row_style(
                    sheet=sheet,
                    source_row=last_row,
                    target_row=new_row,
                    start_column=2,
                    end_column=12,
                )

            # =====================================
            # 値の書き込み
            # =====================================

            # B列：No
            sheet[
                f"B{new_row}"
            ] = new_no

            # C列：部署名
            sheet[
                f"C{new_row}"
            ] = data.department

            # D列：見積／請求番号
            sheet[
                f"D{new_row}"
            ] = new_estimate_no

            # F列：業務委託計画書No
            # 番号がない帳票では空欄にする。
            sheet[
                f"F{new_row}"
            ] = (
                application_number
                if application_number is not None
                else "-"
            )

            # G列：車種コード
            sheet[
                f"G{new_row}"
            ] = data.model_code

            # H列：委託金額
            sheet[
                f"H{new_row}"
            ] = amount_value

            # I列：見積月
            sheet[
                f"I{new_row}"
            ] = estimate_month

            # J列：発行日
            sheet[
                f"J{new_row}"
            ] = today.date()

            # K列：発行者
            sheet[
                f"K{new_row}"
            ] = user_name

            # =====================================
            # フォント設定
            # B列～L列
            # =====================================
            for column in range(
                2,
                13,
            ):
                cell = sheet.cell(
                    row=new_row,
                    column=column,
                )

                current_font = cell.font

                cell.font = Font(
                    name="Meiryo UI",
                    size=10,
                    bold=current_font.bold,
                    italic=current_font.italic,
                    vertAlign=(
                        current_font.vertAlign
                    ),
                    underline=(
                        current_font.underline
                    ),
                    strike=current_font.strike,
                    color=copy(
                        current_font.color
                    ),
                )

            # =====================================
            # 配置設定
            # =====================================

            # C列・G列は前行の左右配置を維持
            for column_letter in (
                "C",
                "G",
            ):
                cell = sheet[
                    f"{column_letter}{new_row}"
                ]

                current_alignment = (
                    cell.alignment
                )

                cell.alignment = Alignment(
                    horizontal=(
                        current_alignment.horizontal
                    ),
                    vertical="center",
                    wrap_text=(
                        current_alignment.wrap_text
                    ),
                    shrink_to_fit=(
                        current_alignment.shrink_to_fit
                    ),
                    text_rotation=(
                        current_alignment.text_rotation
                    ),
                )

            # その他の入力セルは上下・左右中央揃え
            for column_letter in (
                "B",
                "D",
                "F",
                "H",
                "I",
                "J",
                "K",
            ):
                cell = sheet[
                    f"{column_letter}{new_row}"
                ]

                current_alignment = (
                    cell.alignment
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=(
                        current_alignment.wrap_text
                    ),
                    shrink_to_fit=(
                        current_alignment.shrink_to_fit
                    ),
                    text_rotation=(
                        current_alignment.text_rotation
                    ),
                )

            # =====================================
            # 数値・日付の表示形式
            # =====================================

            # F列：業務委託計画書Noがある場合のみ表示形式を設定
            if application_number is not None:
                sheet[
                    f"F{new_row}"
                ].number_format = '"ITK"0'
            else:
                sheet[
                    f"F{new_row}"
                ].number_format = "@"

            # H列：3桁カンマ区切り
            sheet[
                f"H{new_row}"
            ].number_format = '#,##0'

            # J列：日付のみ表示
            sheet[
                f"J{new_row}"
            ].number_format = "'yy/m/d"

            # =====================================
            # B列幅を内容に合わせて調整
            # =====================================
            self.adjust_column_width(
                sheet=sheet,
                column_letter="B",
                minimum_width=4,
                maximum_width=12,
            )

            # =====================================
            # 保存
            # =====================================
            workbook.save(
                excel_path
            )

            return {
                "sheet_name": sheet.title,
                "row": new_row,
                "used_preassigned_row": (
                    used_preassigned_row
                ),
                "no": str(new_no),
                "estimate_no": str(
                    new_estimate_no
                ),
                "issue_date": (
                    today.strftime(
                        "%Y/%m/%d"
                    )
                ),
                "issuer_name": user_name,
                "user_name": user_name,
                "amount": amount_value,
                "application_no": (
                    application_number
                    if application_number is not None
                    else "-"
                ),
            }

        finally:
            workbook.close()

    # =====================================
    # 全シートのF列から重複確認
    # =====================================
    def find_duplicate_application_no(
        self,
        workbook,
        application_no: int,
    ):
        """
        管理台帳の全シートについて、
        F列に同じ業務計画書Noがないか検索する。

        Excelのセル値が次のいずれでも
        同じ番号として比較する。

            14642
            ITK14642
            14,642
            14642.0

        Returns:
            重複あり:
                (該当シート, 行番号)

            重複なし:
                None
        """

        if application_no <= 0:
            return None

        for sheet in workbook.worksheets:
            duplicate_row = (
                self.find_application_no_row(
                    sheet=sheet,
                    application_no=application_no,
                )
            )

            if duplicate_row is not None:
                return (
                    sheet,
                    duplicate_row,
                )

        return None

    # =====================================
    # 指定シートのF列から業務計画書No検索
    # =====================================
    def find_application_no_row(
        self,
        sheet,
        application_no: int,
    ) -> int | None:
        """
        指定シートのF列から、
        業務計画書Noを検索する。

        一致した場合は行番号を返す。
        一致しない場合はNoneを返す。
        """

        if application_no <= 0:
            return None

        for row_number in range(
            1,
            sheet.max_row + 1,
        ):
            cell_value = sheet[
                f"F{row_number}"
            ].value

            if cell_value is None:
                continue

            registered_number = (
                self.to_int(
                    cell_value
                )
            )

            if registered_number <= 0:
                continue

            if (
                registered_number
                == application_no
            ):
                return row_number

        return None

    # =====================================
    # 全シートのC列から依頼部署を検索
    # =====================================
    def find_sheet_by_department(
        self,
        workbook,
        department: str,
    ):
        """
        全シートのC列を検索し、
        依頼部署と一致する値があるシートを返す。
        """

        normalized_department = (
            self.normalize_text(
                department
            )
        )

        if not normalized_department:
            return None

        for sheet in workbook.worksheets:

            for cell in sheet["C"]:

                if cell.value is None:
                    continue

                normalized_value = (
                    self.normalize_text(
                        cell.value
                    )
                )

                if not normalized_value:
                    continue

                if (
                    normalized_value
                    == normalized_department
                    or normalized_department
                    in normalized_value
                    or normalized_value
                    in normalized_department
                ):
                    return sheet

        return None

    # =====================================
    # 文字列正規化
    # =====================================
    def normalize_text(
        self,
        value,
    ) -> str:
        """
        検索比較用に文字列を正規化する。

        ・全角英数字を半角へ統一
        ・前後空白を除去
        ・半角・全角スペースを除去
        """

        if value is None:
            return ""

        normalized = unicodedata.normalize(
            "NFKC",
            str(value),
        )

        normalized = (
            normalized
            .strip()
            .replace("　", "")
            .replace(" ", "")
        )

        return normalized

    # =====================================
    # 採番済みの未使用行を取得
    # =====================================
    def find_available_row(
        self,
        sheet,
    ) -> int | None:
        """
        B列とD列に番号が用意されており、
        F列が空欄の最初の行を返す。

        F列には使用済み案件の場合、
        業務委託計画書Noまたは「-」が入るため、
        空欄の行を未使用として扱う。
        """

        for row_number in range(
            2,
            sheet.max_row + 1,
        ):
            no_value = sheet[
                f"B{row_number}"
            ].value

            estimate_no_value = sheet[
                f"D{row_number}"
            ].value

            application_no_value = sheet[
                f"F{row_number}"
            ].value

            no_number = self.to_int(
                no_value
            )

            estimate_no_number = self.to_int(
                estimate_no_value
            )

            application_no_text = str(
                application_no_value
                or ""
            ).strip()

            if no_number <= 0:
                continue

            if estimate_no_number <= 0:
                continue

            if application_no_text:
                continue

            return row_number

        return None

    # =====================================
    # B列の最終データ行を取得
    # =====================================
    def find_last_row(
        self,
        sheet,
    ) -> int:
        """
        B列を下から検索し、
        最後に値が入力されている行を返す。
        """

        for row_number in range(
            sheet.max_row,
            1,
            -1,
        ):
            value = sheet[
                f"B{row_number}"
            ].value

            if value is None:
                continue

            if str(value).strip():
                return row_number

        return 1

    # =====================================
    # 数値変換
    # =====================================
    def to_int(
        self,
        value,
    ) -> int:
        """
        値から数字部分を取り出し、
        intへ変換する。

        例:
            ITK14642 → 14642
            16,107,097 → 16107097
            14642.0 → 14642
        """

        if value is None:
            return 0

        if isinstance(
            value,
            bool,
        ):
            return 0

        if isinstance(
            value,
            int,
        ):
            return value

        if isinstance(
            value,
            float,
        ):
            return int(value)

        text = unicodedata.normalize(
            "NFKC",
            str(value),
        ).strip()

        text_without_commas = (
            text.replace(",", "")
        )

        try:
            return int(
                float(
                    text_without_commas
                )
            )

        except (
            TypeError,
            ValueError,
        ):
            digits = "".join(
                character
                for character in text
                if character.isdigit()
            )

            return (
                int(digits)
                if digits
                else 0
            )

    # =====================================
    # 委託金額の変換
    # =====================================
    def normalize_amount(
        self,
        value,
    ) -> int | str:
        """
        委託金額をExcelへ書き込める形に整える。

        例:
            ￥924,000
            ¥924,000
            924,000円
            924000
            924000.0

        戻り値:
            924000

        数値化できない場合は、
        元の文字列を返す。
        """

        if value is None:
            return ""

        if isinstance(
            value,
            bool,
        ):
            return ""

        if isinstance(
            value,
            int,
        ):
            return value

        if isinstance(
            value,
            float,
        ):
            return int(value)

        text = unicodedata.normalize(
            "NFKC",
            str(value),
        )

        text = (
            text
            .strip()
            .replace("￥", "")
            .replace("¥", "")
            .replace("\\", "")
            .replace(",", "")
            .replace("円", "")
            .replace("税込", "")
            .replace("税抜", "")
            .replace(" ", "")
            .replace("　", "")
        )

        try:
            return int(
                float(text)
            )

        except (
            TypeError,
            ValueError,
        ):
            return str(
                value
            ).strip()

    # =====================================
    # 翌月を「8月」形式で返す
    # =====================================
    def get_next_month_text(
        self,
        date_value: datetime,
    ) -> str:
        """
        指定日の翌月を「8月」の形式で返す。
        """

        next_month = (
            date_value.month + 1
        )

        if next_month == 13:
            next_month = 1

        return f"{next_month}月"

    # =====================================
    # 前行の書式を新しい行へコピー
    # =====================================
    def copy_row_style(
        self,
        sheet,
        source_row: int,
        target_row: int,
        start_column: int,
        end_column: int,
    ) -> None:
        """
        前行のフォント、塗りつぶし、罫線、
        配置、表示形式、保護、行の高さを
        新しい行へコピーする。
        """

        source_height = (
            sheet.row_dimensions[
                source_row
            ].height
        )

        if source_height is not None:
            sheet.row_dimensions[
                target_row
            ].height = source_height

        for column in range(
            start_column,
            end_column + 1,
        ):
            source_cell = sheet.cell(
                row=source_row,
                column=column,
            )

            target_cell = sheet.cell(
                row=target_row,
                column=column,
            )

            if not source_cell.has_style:
                continue

            target_cell.font = copy(
                source_cell.font
            )

            target_cell.fill = copy(
                source_cell.fill
            )

            target_cell.border = copy(
                source_cell.border
            )

            target_cell.alignment = copy(
                source_cell.alignment
            )

            target_cell.number_format = (
                source_cell.number_format
            )

            target_cell.protection = copy(
                source_cell.protection
            )

    # =====================================
    # 列幅調整
    # =====================================
    def adjust_column_width(
        self,
        sheet,
        column_letter: str,
        minimum_width: float = 4,
        maximum_width: float = 50,
    ) -> None:
        """
        指定列の内容に合わせて列幅を調整する。
        """

        maximum_length = 0

        for cell in sheet[
            column_letter
        ]:
            if cell.value is None:
                continue

            text_length = len(
                str(cell.value)
            )

            if (
                text_length
                > maximum_length
            ):
                maximum_length = (
                    text_length
                )

        calculated_width = (
            maximum_length + 2
        )

        calculated_width = max(
            minimum_width,
            calculated_width,
        )

        calculated_width = min(
            maximum_width,
            calculated_width,
        )

        sheet.column_dimensions[
            column_letter
        ].width = calculated_width