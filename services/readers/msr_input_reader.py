import os
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl
import xlrd


@dataclass
class MsrRequestRow:
    """見積書発行依頼の明細1行分。"""

    request_no: str
    construction_name: str
    order_period: str
    department: str
    amount: float


@dataclass
class MsrRequest:
    """見積書発行依頼1ファイル分の読み取り結果。"""

    source_path: str
    department: str
    request_year: int
    # 依頼元の上段（例：豊田事業部）。台帳の部署名用
    department_upper: str = ""
    rows: list = field(default_factory=list)


class MsrInputReader:
    """
    MSRの見積書発行依頼Excel（.xls／.xlsx）を読み取る。

    フォルダ内から自動判別するため、
    見積書発行依頼のレイアウトでないファイルは
    parse時に例外を送出する。
    """

    SHEET_NAME = "見積書発行依頼"

    # データ行はExcelの17行目から（0始まりで16）
    DATA_START_INDEX = 16
    DATA_END_INDEX = 29

    # =====================================
    # 解析
    # =====================================
    def parse(self, path: str) -> MsrRequest:

        ext = os.path.splitext(path)[1].lower()

        if ext == ".xls":
            return self._parse_xls(path)

        if ext in (".xlsx", ".xlsm"):
            return self._parse_xlsx(path)

        raise ValueError(
            "Excelファイルではありません。"
        )

    # =====================================
    # .xls読み取り
    # =====================================
    def _parse_xls(self, path: str) -> MsrRequest:

        book = xlrd.open_workbook(path)

        if self.SHEET_NAME in book.sheet_names():
            sheet = book.sheet_by_name(self.SHEET_NAME)
        else:
            sheet = book.sheet_by_index(0)

        def cell(row, col):
            if row >= sheet.nrows or col >= sheet.ncols:
                return None
            value = sheet.cell_value(row, col)
            return value if value != "" else None

        self._validate_title(cell(2, 0))

        # 依頼元①（I7＋I8）
        department_upper = str(cell(6, 8) or "").strip()

        department = self._join_department(
            cell(6, 8),
            cell(7, 8),
        )

        # 依頼日（I1）から年を取得
        request_year = datetime.today().year

        i1 = cell(0, 8)

        if isinstance(i1, float):
            request_year = xlrd.xldate_as_datetime(
                i1,
                book.datemode,
            ).year

        rows = []

        for r in range(
            self.DATA_START_INDEX,
            self.DATA_END_INDEX + 1,
        ):
            request_no = cell(r, 1)

            if not request_no:
                continue

            rows.append(
                MsrRequestRow(
                    request_no=str(request_no).strip(),
                    construction_name=str(
                        cell(r, 3) or ""
                    ).strip(),
                    order_period=str(
                        cell(r, 6) or ""
                    ).strip(),
                    department=str(
                        cell(r, 7) or ""
                    ).strip(),
                    amount=float(cell(r, 8) or 0),
                )
            )

        return self._build_request(
            path=path,
            department=department,
            department_upper=department_upper,
            request_year=request_year,
            rows=rows,
        )

    # =====================================
    # .xlsx読み取り
    # =====================================
    def _parse_xlsx(self, path: str) -> MsrRequest:

        book = openpyxl.load_workbook(
            path,
            data_only=True,
        )

        if self.SHEET_NAME in book.sheetnames:
            sheet = book[self.SHEET_NAME]
        else:
            sheet = book[book.sheetnames[0]]

        self._validate_title(sheet["A3"].value)

        department_upper = str(
            sheet["I7"].value or ""
        ).strip()

        department = self._join_department(
            sheet["I7"].value,
            sheet["I8"].value,
        )

        request_year = datetime.today().year

        i1 = sheet["I1"].value

        if isinstance(i1, datetime):
            request_year = i1.year

        rows = []

        for r in range(
            self.DATA_START_INDEX + 1,
            self.DATA_END_INDEX + 2,
        ):
            request_no = sheet.cell(row=r, column=2).value

            if not request_no:
                continue

            rows.append(
                MsrRequestRow(
                    request_no=str(request_no).strip(),
                    construction_name=str(
                        sheet.cell(row=r, column=4).value
                        or ""
                    ).strip(),
                    order_period=str(
                        sheet.cell(row=r, column=7).value
                        or ""
                    ).strip(),
                    department=str(
                        sheet.cell(row=r, column=8).value
                        or ""
                    ).strip(),
                    amount=float(
                        sheet.cell(row=r, column=9).value
                        or 0
                    ),
                )
            )

        return self._build_request(
            path=path,
            department=department,
            department_upper=department_upper,
            request_year=request_year,
            rows=rows,
        )

    # =====================================
    # タイトル検証（自動判別用）
    # =====================================
    def _validate_title(self, value):

        title = str(value or "")

        if "見積書発行依頼" not in title:
            raise ValueError(
                "見積書発行依頼のフォーマットではありません。"
            )

    # =====================================
    # 依頼元①の結合
    # =====================================
    def _join_department(self, upper, lower):

        parts = [
            str(x).strip()
            for x in (upper, lower)
            if x is not None and str(x).strip()
        ]

        return " ".join(parts)

    # =====================================
    # 読み取り結果の組み立て
    # =====================================
    def _build_request(
        self,
        path: str,
        department: str,
        department_upper: str,
        request_year: int,
        rows: list,
    ) -> MsrRequest:

        if not rows:
            raise ValueError(
                "明細行が見つかりません。"
            )

        return MsrRequest(
            source_path=path,
            department=department,
            department_upper=department_upper,
            request_year=request_year,
            rows=rows,
        )
