import re
import unicodedata
from collections.abc import Iterable

from openpyxl import load_workbook


class UserMasterReader:
    """
    利用者一覧Excelから、メールアドレスに対応する
    利用者名を取得する。

    実際の列構成：
        B列：No
        C列：利用者名
        D列：メールアドレス
    """

    def find_user_name(
        self,
        excel_path: str,
        email_addresses: Iterable[str],
    ) -> str:
        """
        複数のメールアドレス候補を、
        全シートのD列から検索する。

        一致した行のC列にある利用者名を返す。
        """

        target_emails = {
            normalized_email
            for email in email_addresses
            if (
                normalized_email
                := self.normalize_email(email)
            )
        }

        if not target_emails:
            return ""

        workbook = load_workbook(
            excel_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )

        try:
            for sheet in workbook.worksheets:

                for row_number in range(
                    1,
                    sheet.max_row + 1,
                ):
                    # C列：利用者名
                    name_cell = sheet[
                        f"C{row_number}"
                    ]

                    # D列：メールアドレス
                    email_cell = sheet[
                        f"D{row_number}"
                    ]

                    email_candidates = (
                        self.get_cell_email_candidates(
                            email_cell
                        )
                    )

                    for candidate in email_candidates:
                        normalized_candidate = (
                            self.normalize_email(
                                candidate
                            )
                        )

                        if (
                            normalized_candidate
                            not in target_emails
                        ):
                            continue

                        if name_cell.value is None:
                            return ""

                        return str(
                            name_cell.value
                        ).strip()

            return ""

        finally:
            workbook.close()

    def get_cell_email_candidates(
        self,
        cell,
    ) -> list[str]:
        """
        セルからメールアドレス候補を取得する。

        対応対象：
        ・セルに表示されている値
        ・ハイパーリンクのリンク先
        ・HYPERLINK関数
        """

        candidates = []

        # セルの表示値
        if cell.value is not None:
            candidates.append(
                str(cell.value)
            )

        # Excelのハイパーリンク
        if cell.hyperlink:
            target = getattr(
                cell.hyperlink,
                "target",
                None,
            )

            if target:
                candidates.append(
                    str(target)
                )

            display = getattr(
                cell.hyperlink,
                "display",
                None,
            )

            if display:
                candidates.append(
                    str(display)
                )

        # HYPERLINK関数
        if isinstance(
            cell.value,
            str,
        ):
            formula_matches = re.findall(
                r'["\'](?:mailto:)?'
                r'([^"\']+@[^"\']+)["\']',
                cell.value,
                flags=re.IGNORECASE,
            )

            candidates.extend(
                formula_matches
            )

        return candidates

    def normalize_email(
        self,
        value,
    ) -> str:
        """
        メールアドレスを比較用に正規化する。
        """

        if value is None:
            return ""

        text = unicodedata.normalize(
            "NFKC",
            str(value),
        )

        text = text.strip().lower()

        if text.startswith("mailto:"):
            text = text[
                len("mailto:"):
            ]

        # URLパラメータ除去
        text = text.split("?")[0]

        # ゼロ幅文字除去
        for character in (
            "\u200b",
            "\u200c",
            "\u200d",
            "\ufeff",
        ):
            text = text.replace(
                character,
                "",
            )

        # 空白・改行除去
        text = re.sub(
            r"\s+",
            "",
            text,
        )

        return text