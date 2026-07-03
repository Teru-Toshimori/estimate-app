from openpyxl import load_workbook
from models.estimate_data import EstimateData


class LedgerReader:

    def find(self, excel_path: str, application_no: str) -> EstimateData:
        """
        管理台帳から申請書Noを検索し、
        見積番号(D列)・発行日(J列)を取得する
        """

        print("=" * 50)
        print("検索開始")

        # PDFの「ITK14642」→ Excelの「14642」
        search_no = application_no.replace("ITK", "").strip()

        print("検索文字列:", search_no)

        workbook = load_workbook(
            excel_path,
            data_only=True
        )

        print("シート一覧:", workbook.sheetnames)

        data = EstimateData()

        # 全シート検索
        for sheet in workbook.worksheets:

            print("検索シート:", sheet.title)

            for row in sheet.iter_rows():

                for cell in row:

                    if cell.value is None:
                        continue

                    value = str(cell.value).strip()

                    if value == search_no:

                        print("一致しました！")
                        print("シート:", sheet.title)
                        print("セル:", cell.coordinate)

                        row_no = cell.row

                        estimate_no = sheet[f"D{row_no}"].value
                        issue_date = sheet[f"J{row_no}"].value

                        print("見積番号:", estimate_no)
                        print("発行日:", issue_date)

                        # 見積番号
                        data.estimate_no = (
                            "" if estimate_no is None else str(estimate_no)
                        )

                        # 発行日
                        if issue_date is None:
                            data.issue_date = ""

                        elif hasattr(issue_date, "strftime"):
                            data.issue_date = issue_date.strftime("%Y/%m/%d")

                        else:
                            data.issue_date = str(issue_date)

                        workbook.close()

                        return data

        print("一致するデータはありませんでした")

        workbook.close()

        return data